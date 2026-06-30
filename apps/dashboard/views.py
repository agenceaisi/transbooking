"""Read-only dashboard endpoints (PROMPT 10).

No models live in this app: every view delegates to ``services.py`` for ORM
aggregations and serialises the result. All dashboards are cached for 5 minutes
(``cache_page(300)``) and varied on the ``Authorization`` header so a cached
response is NEVER served across users — this preserves strict multi-tenant
isolation (cf. CLAUDE.md).
"""
import csv
from datetime import date
from io import BytesIO

from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_headers
from rest_framework.exceptions import NotFound, ValidationError
from rest_framework.negotiation import DefaultContentNegotiation
from rest_framework.response import Response
from rest_framework.views import APIView

from utils.permissions import IsAgent, IsCompanyAdmin, IsSuperAdmin, IsVoyageur

from . import serializers, services

# Cache de 5 minutes pour tous les tableaux de bord (PROMPT 10), isole par JWT.
_DASHBOARD_CACHE = [cache_page(300), vary_on_headers("Authorization")]


def _parse_date(value: str | None) -> date | None:
    """Parse an ISO ``YYYY-MM-DD`` query param, or ``None`` if absent."""
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        raise ValidationError(
            {"date": "Format de date invalide, attendu AAAA-MM-JJ."}
        )


class ExportContentNegotiation(DefaultContentNegotiation):
    """Neutralise le parametre `?format=` reserve par DRF pour l'export.

    L'export reutilise `?format=pdf|excel` pour choisir le type de fichier, ce
    qui entrerait en conflit avec la selection de renderer de DRF (404).
    """

    def select_renderer(self, request, renderers, format_suffix=None):
        return renderers[0], renderers[0].media_type


# --------------------------------------------------------------------------- #
# Voyageur
# --------------------------------------------------------------------------- #
@method_decorator(_DASHBOARD_CACHE, name="get")
class TravelerDashboardView(APIView):
    """GET /api/v1/dashboard/traveler/ — accueil voyageur."""

    permission_classes = [IsVoyageur]
    serializer_class = serializers.TravelerDashboardSerializer

    def get(self, request):
        data = services.traveler_dashboard(request.user)
        return Response(self.serializer_class(data).data)


# --------------------------------------------------------------------------- #
# Agent
# --------------------------------------------------------------------------- #
@method_decorator(_DASHBOARD_CACHE, name="get")
class AgentDashboardView(APIView):
    """GET /api/v1/agent/dashboard/ — tableau de bord de l'agent."""

    permission_classes = [IsAgent]
    serializer_class = serializers.AgentDashboardSerializer

    def get(self, request):
        data = services.agent_dashboard(request.user)
        return Response(self.serializer_class(data).data)


# --------------------------------------------------------------------------- #
# Company admin
# --------------------------------------------------------------------------- #
class CompanyDashboardMixin:
    """Resolution de la compagnie courante et de la periode demandee."""

    permission_classes = [IsCompanyAdmin]

    def get_company(self):
        company = getattr(self.request.user, "administered_company", None)
        if company is None:
            raise NotFound("Aucune compagnie associee a cet utilisateur.")
        return company

    def get_period(self):
        params = self.request.query_params
        try:
            return services.resolve_period(
                params.get("period", "month"),
                _parse_date(params.get("start_date")),
                _parse_date(params.get("end_date")),
            )
        except ValueError as exc:
            raise ValidationError(str(exc))


@method_decorator(_DASHBOARD_CACHE, name="get")
class CompanyDashboardView(CompanyDashboardMixin, APIView):
    """GET /api/v1/company/dashboard/ — KPI de la compagnie avec deltas."""

    serializer_class = serializers.CompanyOverviewSerializer

    def get(self, request):
        data = services.company_overview(self.get_company(), self.get_period())
        return Response(self.serializer_class(data).data)


@method_decorator(_DASHBOARD_CACHE, name="get")
class CompanyRevenueChartView(CompanyDashboardMixin, APIView):
    """GET /api/v1/company/dashboard/revenue-chart/ — recettes par jour/semaine."""

    serializer_class = serializers.RevenuePointSerializer

    def get(self, request):
        data = services.company_revenue_chart(
            self.get_company(), self.get_period()
        )
        return Response(self.serializer_class(data, many=True).data)


@method_decorator(_DASHBOARD_CACHE, name="get")
class CompanyFillRateByRouteView(CompanyDashboardMixin, APIView):
    """GET /api/v1/company/dashboard/fill-rate-by-route/ — taux de remplissage."""

    serializer_class = serializers.FillRateByRouteSerializer

    def get(self, request):
        data = services.company_fill_rate_by_route(
            self.get_company(), self.get_period()
        )
        return Response(self.serializer_class(data, many=True).data)


@method_decorator(_DASHBOARD_CACHE, name="get")
class CompanyPaymentBreakdownView(CompanyDashboardMixin, APIView):
    """GET /api/v1/company/dashboard/payment-breakdown/ — repartition paiements."""

    serializer_class = serializers.PaymentBreakdownSerializer

    def get(self, request):
        data = services.company_payment_breakdown(
            self.get_company(), self.get_period()
        )
        return Response(self.serializer_class(data, many=True).data)


@method_decorator(_DASHBOARD_CACHE, name="get")
class CompanyTopRoutesView(CompanyDashboardMixin, APIView):
    """GET /api/v1/company/dashboard/top-routes/ — top 5 trajets par recette."""

    serializer_class = serializers.TopRouteSerializer

    def get(self, request):
        data = services.company_top_routes(self.get_company(), self.get_period())
        return Response(self.serializer_class(data, many=True).data)


@method_decorator(_DASHBOARD_CACHE, name="get")
class CompanyAgentActivityView(CompanyDashboardMixin, APIView):
    """GET /api/v1/company/dashboard/agent-activity/ — activite du jour par agent."""

    serializer_class = serializers.AgentActivitySerializer

    def get(self, request):
        data = services.company_agent_activity(self.get_company())
        return Response(self.serializer_class(data, many=True).data)


@method_decorator(_DASHBOARD_CACHE, name="get")
class CompanyAlertsView(CompanyDashboardMixin, APIView):
    """GET /api/v1/company/dashboard/alerts/ — compteurs d'alertes operationnelles."""

    serializer_class = serializers.CompanyAlertsSerializer

    def get(self, request):
        data = services.company_alerts(self.get_company())
        return Response(self.serializer_class(data).data)


class CompanyDashboardExportView(CompanyDashboardMixin, APIView):
    """GET /api/v1/company/dashboard/export/?format=pdf|excel — rapport complet.

    Non mis en cache : renvoie un fichier (xlsx/pdf) genere a la volee.
    """

    content_negotiation_class = ExportContentNegotiation
    serializer_class = serializers.CompanyOverviewSerializer

    def get(self, request):
        company = self.get_company()
        period = self.get_period()
        overview = services.company_overview(company, period)
        top_routes = services.company_top_routes(company, period)
        breakdown = services.company_payment_breakdown(company, period)
        export_format = request.query_params.get("format", "excel").lower()
        if export_format == "pdf":
            return self._export_pdf(company, overview, top_routes, breakdown)
        return self._export_excel(company, overview, top_routes, breakdown)

    def _export_excel(self, company, overview, top_routes, breakdown):
        try:
            from openpyxl import Workbook
        except ImportError:
            return self._export_csv(company, overview, top_routes, breakdown)

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Synthese"
        summary.append(["Compagnie", company.name])
        summary.append(["Periode", overview["period"]])
        summary.append(["Recette totale (FCFA)", overview["revenue_total"]])
        summary.append(["Taux de remplissage moyen (%)", overview["fill_rate_avg"]])
        summary.append(["Reservations", overview["bookings_count"]])
        summary.append(["Note moyenne", overview["avg_rating"]])

        routes_sheet = workbook.create_sheet("Top trajets")
        routes_sheet.append(["Trajet", "Recette (FCFA)", "Passagers"])
        for row in top_routes:
            routes_sheet.append([row["route"], row["revenue"], row["passengers"]])

        pay_sheet = workbook.create_sheet("Paiements")
        pay_sheet.append(["Methode", "Montant (FCFA)", "Part (%)"])
        for row in breakdown:
            pay_sheet.append([row["method"], row["amount"], row["pct"]])

        buffer = BytesIO()
        workbook.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            'attachment; filename="dashboard.xlsx"'
        )
        return response

    def _export_csv(self, company, overview, top_routes, breakdown):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="dashboard.csv"'
        writer = csv.writer(response)
        writer.writerow(["Compagnie", company.name])
        writer.writerow(["Periode", overview["period"]])
        writer.writerow(["Recette totale (FCFA)", overview["revenue_total"]])
        writer.writerow(["Taux de remplissage moyen (%)", overview["fill_rate_avg"]])
        writer.writerow(["Reservations", overview["bookings_count"]])
        writer.writerow(["Note moyenne", overview["avg_rating"]])
        writer.writerow([])
        writer.writerow(["Trajet", "Recette (FCFA)", "Passagers"])
        for row in top_routes:
            writer.writerow([row["route"], row["revenue"], row["passengers"]])
        return response

    def _export_pdf(self, company, overview, top_routes, breakdown):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas

        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        _, height = A4
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(20 * mm, height - 20 * mm, f"Tableau de bord — {company.name}")
        pdf.setFont("Helvetica", 9)
        y = height - 30 * mm
        lines = [
            f"Periode : {overview['period']}",
            f"Recette totale : {overview['revenue_total']} FCFA",
            f"Taux de remplissage moyen : {overview['fill_rate_avg']} %",
            f"Reservations : {overview['bookings_count']}",
            f"Note moyenne : {overview['avg_rating']}",
            "",
            "Top trajets :",
        ]
        for row in top_routes:
            lines.append(
                f"  {row['route']} — {row['revenue']} FCFA "
                f"({row['passengers']} passagers)"
            )
        for line in lines:
            pdf.drawString(20 * mm, y, line)
            y -= 7 * mm
            if y < 20 * mm:
                pdf.showPage()
                pdf.setFont("Helvetica", 9)
                y = height - 20 * mm
        pdf.showPage()
        pdf.save()
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="dashboard.pdf"'
        return response


# --------------------------------------------------------------------------- #
# Super admin
# --------------------------------------------------------------------------- #
@method_decorator(_DASHBOARD_CACHE, name="get")
class SuperDashboardView(APIView):
    """GET /api/v1/super/dashboard/ — vue plateforme globale."""

    permission_classes = [IsSuperAdmin]
    serializer_class = serializers.SuperOverviewSerializer

    def get(self, request):
        data = services.super_overview()
        return Response(self.serializer_class(data).data)


@method_decorator(_DASHBOARD_CACHE, name="get")
class SuperRevenueByCompanyView(APIView):
    """GET /api/v1/super/dashboard/revenue-by-company/ — recettes par compagnie."""

    permission_classes = [IsSuperAdmin]
    serializer_class = serializers.RevenueByCompanySerializer

    def get(self, request):
        data = services.super_revenue_by_company()
        return Response(self.serializer_class(data, many=True).data)


@method_decorator(_DASHBOARD_CACHE, name="get")
class SuperBookingsChartView(APIView):
    """GET /api/v1/super/dashboard/bookings-chart/ — reservations globales dans le temps."""

    permission_classes = [IsSuperAdmin]
    serializer_class = serializers.BookingsChartPointSerializer

    def get(self, request):
        try:
            period = services.resolve_period(
                request.query_params.get("period", "month"),
                _parse_date(request.query_params.get("start_date")),
                _parse_date(request.query_params.get("end_date")),
            )
        except ValueError as exc:
            raise ValidationError(str(exc))
        data = services.super_bookings_chart(period)
        return Response(self.serializer_class(data, many=True).data)
